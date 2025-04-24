import streamlit as st
import pandas as pd
import datetime
import openpyxl
from io import BytesIO
from fpdf import FPDF
import matplotlib.pyplot as plt
import matplotlib as mpl
import tax_engine

# Set up a professional chart theme
def set_chart_theme():
    # Set the font family
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'DejaVu Sans']
    
    # Set the font sizes
    plt.rcParams['font.size'] = 10
    plt.rcParams['axes.titlesize'] = 14
    plt.rcParams['axes.labelsize'] = 12
    plt.rcParams['xtick.labelsize'] = 10
    plt.rcParams['ytick.labelsize'] = 10
    plt.rcParams['legend.fontsize'] = 10
    
    # Set the colors
    plt.rcParams['axes.prop_cycle'] = plt.cycler(color=['#4e79a7', '#f28e2c', '#59a14f', '#e15759', '#76b7b2', '#edc949', '#b07aa1', '#ff9da7'])
    
    # Set the grid style
    plt.rcParams['axes.grid'] = True
    plt.rcParams['grid.alpha'] = 0.3
    plt.rcParams['grid.linestyle'] = '--'
    
    # Set the figure size and DPI
    plt.rcParams['figure.figsize'] = (10, 6)
    plt.rcParams['figure.dpi'] = 100
    
    # Set the background color
    plt.rcParams['figure.facecolor'] = '#f8f9fa'
    plt.rcParams['axes.facecolor'] = '#f8f9fa'
    
    # Set the spine visibility
    plt.rcParams['axes.spines.top'] = False
    plt.rcParams['axes.spines.right'] = False

# Apply the chart theme
set_chart_theme()

st.set_page_config(page_title="RSU Manager", layout="wide")
st.title("📊 RSU Management Dashboard")

# Sidebar inputs
st.sidebar.markdown("This is a demo app and not provide any financial advise.All data is stored in the browser session only—nothing is sent to a server.")
st.sidebar.header("📝 RSU Input")
company = st.sidebar.text_input("Company Name", "TechCorp")
tax_rate = st.sidebar.slider("Marginal Tax Rate (%)", 0, 50, 45)

# Simplified tax settings
tax_residency = "Australian Resident"  # Default value

# Load sample data
if st.sidebar.button("📂 Load Sample Data"):
    st.session_state["load_sample"] = True
    
    # Load vesting data
    vesting_data = pd.read_excel("sample_rsu_data.xlsx")
    
    # Print original data for debugging
    print("Original vesting data:")
    print(vesting_data)
    print("Columns:", vesting_data.columns.tolist())
    
    # Process vesting data with tax engine
    vesting_data = tax_engine.process_vesting_data(vesting_data, tax_rate)
    
    # Print processed data for debugging
    print("Processed vesting data:")
    print(vesting_data)
    print("Columns:", vesting_data.columns.tolist())
    
    st.session_state["vesting_data"] = vesting_data
    
    # Load sales data
    sales_data = pd.read_excel("sample_rsu_sales.xlsx")
    # Process sales data with tax engine
    sales_data = tax_engine.process_sales_data(sales_data, tax_rate)
    st.session_state["sales_data"] = sales_data
    
    st.sidebar.success("Sample data loaded successfully!")

# Upload data
vesting_file = st.sidebar.file_uploader("Upload Vesting Schedule (Excel)", type=["xlsx"], key="vesting")
sales_file = st.sidebar.file_uploader("Upload RSU Sales Schedule (Excel)", type=["xlsx"], key="sales")


# Add Buy Me a Coffee section
st.sidebar.markdown("### ☕ Support the Developer")
st.sidebar.markdown("""
<a href="https://buymeacoffee.com/cloudnamaste" target="_blank">
    <img src="https://cdn.buymeacoffee.com/buttons/v2/default-yellow.png" 
                alt="Buy Me A Coffee" 
                width="200">
</a>
""", unsafe_allow_html=True)
st.sidebar.markdown("Support this tool's development and maintenance!")


# Tabs
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📆 Vesting Schedule",
    "💸 RSU Sell Tracker",
    "📊 Summary",
    "📝 Tax Return Assistant",
    "💡 Optimization"
])

# -------- TAB 1: VESTING --------
with tab1:
    if "vesting_data" in st.session_state:
        schedule = st.session_state["vesting_data"]
        
        # Display vesting schedule (editable)
        st.subheader(f"Vesting Schedule for {company}")
        
        # Make the dataframe editable
        edited_schedule = st.data_editor(
            schedule.style.format({
                "Gross Value": "${:,.2f}",
                "Tax Payable": "${:,.2f}",
                "Net Value": "${:,.2f}",
                "Marginal Tax Rate": "{:.1f}%"
            }),
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key="vesting_data_editor"
        )
        
        # Process the edited data only if it has changed
        if edited_schedule is not None and not edited_schedule.equals(schedule):
            # Process the edited data with tax engine
            processed_schedule = tax_engine.process_vesting_data(edited_schedule, tax_rate)
            st.session_state["vesting_data"] = processed_schedule
            schedule = processed_schedule
            
            # Force rerun to update the UI immediately with recalculated values
            st.rerun()
        
        # Create visualization for total vest schedule by year
        if not schedule.empty and "Financial Year" in schedule.columns:
            st.markdown("### 📊 Vesting Schedule Visualization")
            
            # Group by Financial Year
            fy_vesting = schedule.groupby("Financial Year")[["RSU Vested", "Gross Value", "Tax Payable", "Net Value"]].sum().reset_index()
            
            # Create bar chart
            fig, ax = plt.subplots()
            
            # Set width of bars
            barWidth = 0.25
            
            # Set position of bars on X axis
            r1 = range(len(fy_vesting))
            r2 = [x + barWidth for x in r1]
            r3 = [x + barWidth for x in r2]
            
            # Create bars with theme colors
            ax.bar(r1, fy_vesting['Gross Value'], width=barWidth, label='Gross Value', color='#4e79a7')
            ax.bar(r2, fy_vesting['Tax Payable'], width=barWidth, label='Tax Payable', color='#e15759')
            ax.bar(r3, fy_vesting['Net Value'], width=barWidth, label='Net Value', color='#59a14f')
            
            # Add labels and title
            ax.set_xlabel('Financial Year')
            ax.set_ylabel('Amount ($)')
            ax.set_title('RSU Vesting Schedule by Financial Year', fontweight='bold')
            ax.set_xticks([r + barWidth for r in range(len(fy_vesting))])
            ax.set_xticklabels(fy_vesting['Financial Year'])
            
            # Add value labels on top of bars
            for i, v in enumerate(fy_vesting['Gross Value']):
                ax.text(i, v + 0.1, f"${v:,.2f}", ha='center', fontsize=9)
            
            for i, v in enumerate(fy_vesting['Tax Payable']):
                ax.text(i + barWidth, v + 0.1, f"${v:,.2f}", ha='center', fontsize=9)
                
            for i, v in enumerate(fy_vesting['Net Value']):
                ax.text(i + 2*barWidth, v + 0.1, f"${v:,.2f}", ha='center', fontsize=9)
            
            # Add legend with better positioning
            ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=3)
            
            # Adjust layout to make room for legend
            plt.tight_layout(rect=[0, 0.05, 1, 0.95])
            
            st.pyplot(fig)
            
            # Add a second chart for RSUs vested by year
            st.markdown("### 📈 RSUs Vested by Year")
            fig2, ax2 = plt.subplots()
            bars = ax2.bar(fy_vesting['Financial Year'], fy_vesting['RSU Vested'], color='#b07aa1')
            ax2.set_xlabel('Financial Year')
            ax2.set_ylabel('Number of RSUs')
            ax2.set_title('Total RSUs Vested by Financial Year', fontweight='bold')
            
            # Add value labels on top of bars
            for bar in bars:
                height = bar.get_height()
                ax2.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                        f'{height:,.0f}', ha='center', va='bottom', fontsize=9)
            
            # Adjust layout
            plt.tight_layout()
            
            st.pyplot(fig2)
        
        # Export to Excel
        
        # Export to Excel
        vesting_buffer = BytesIO()
        with pd.ExcelWriter(vesting_buffer, engine='openpyxl') as writer:
            schedule.to_excel(writer, index=False, sheet_name="Vesting Schedule")
        vesting_buffer.seek(0)
        
        st.download_button("📥 Download Vesting Schedule (Excel)",
                          data=vesting_buffer,
                          file_name="vesting_schedule.xlsx",
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Add button to clear all data
        if st.button("🗑️ Clear All Vesting Data"):
            if st.session_state.get("confirm_clear_vesting", False):
                st.session_state["vesting_data"] = pd.DataFrame()
                st.rerun()
            else:
                st.session_state["confirm_clear_vesting"] = True
                st.warning("Click again to confirm clearing all vesting data.")
    elif vesting_file:
        schedule = pd.read_excel(vesting_file)
        # Process vesting data using tax engine
        schedule = tax_engine.process_vesting_data(schedule, tax_rate)
        
        # Display vesting schedule
        st.subheader(f"Vesting Schedule for {company}")
        st.dataframe(schedule.style.format({
            "Gross Value": "${:,.2f}",
            "Tax Payable": "${:,.2f}",
            "Net Value": "${:,.2f}"
        }))
        
        # Export to Excel
        vesting_buffer = BytesIO()
        with pd.ExcelWriter(vesting_buffer, engine='openpyxl') as writer:
            schedule.to_excel(writer, index=False, sheet_name="Vesting Schedule")
        vesting_buffer.seek(0)
        
        st.download_button("📥 Download Vesting Schedule (Excel)",
                          data=vesting_buffer,
                          file_name="vesting_schedule.xlsx",
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        # Create an empty DataFrame with the necessary columns for vesting data
        empty_vesting_df = pd.DataFrame({
            "Vesting Date": [],
            "RSU Vested": [],
            "FMV at Vesting": [],
            "Marginal Tax Rate": [],
            "Gross Value": [],
            "Tax Payable": [],
            "Net Value": []
        })
        
        # Process the empty DataFrame to ensure it has all required columns
        empty_vesting_df = tax_engine.process_vesting_data(empty_vesting_df, tax_rate)
        
        # Display an editable data grid even when no data is available
        st.info("No vesting data available. You can add entries directly in the grid below.")
        
        # Display formatted data (editable)
        edited_schedule = st.data_editor(
            empty_vesting_df.style.format({
                "Gross Value": "${:,.2f}",
                "Tax Payable": "${:,.2f}",
                "Net Value": "${:,.2f}",
                "Marginal Tax Rate": "{:.1f}%"
            }),
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key="vesting_data_editor_empty"
        )
        
        # Process the edited data only if it has changed and is not empty
        if edited_schedule is not None and not edited_schedule.empty and not edited_schedule.equals(empty_vesting_df):
            # Process the edited data with tax engine
            processed_schedule = tax_engine.process_vesting_data(edited_schedule, tax_rate)
            st.session_state["vesting_data"] = processed_schedule
            
            # Force rerun to update the UI immediately with recalculated values
            st.rerun()

# This section has been moved into the conditional blocks above

# -------- TAB 2: SALES --------
with tab2:
    st.subheader("💸 RSU Sell Tracker & Capital Gains")

    if "sales_data" in st.session_state:
        sell_data = st.session_state["sales_data"]
        
        # Process data
        sell_data = tax_engine.process_sales_data(sell_data, tax_rate)
        
        # No "Add New Sales Entry" button - users can add rows directly in the data editor
        
        # Display formatted data (editable)
        edited_sell_data = st.data_editor(
            sell_data.style.format({
                "Capital Gain": "${:,.2f}",
                "Tax on CG": "${:,.2f}",
                "Net Proceeds": "${:,.2f}",
                "Marginal Tax Rate": "{:.1f}%"
            }),
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key="sales_data_editor_upload"  # Add a unique key for this data editor
        )
        
        # Process the edited data only if it has changed
        if edited_sell_data is not None and not edited_sell_data.equals(sell_data):
            # Process the edited data with tax engine
            processed_sell_data = tax_engine.process_sales_data(edited_sell_data, tax_rate)
            st.session_state["sales_data"] = processed_sell_data
            sell_data = processed_sell_data
            
            # Force rerun to update the UI immediately with recalculated values
            st.rerun()
        
        # Create visualizations
        if not sell_data.empty:
            st.markdown("### 📊 Visualizations")
            
            # Create tabs for different visualizations
            viz_tab1, viz_tab2, viz_tab3 = st.tabs([
                "Capital Gain/Loss by Financial Year",
                "Stock Performance",
                "Capital Gains vs Tax"
            ])
            
            with viz_tab1:
                # New visualization - Capital Gain/Loss by Financial Year
                if "Financial Year" in sell_data.columns:
                    # Group by Financial Year
                    fy_data = sell_data.groupby("Financial Year")["Capital Gain"].sum().reset_index()
                    
                    # Create bar chart
                    fig2, ax2 = plt.subplots()
                    bars = ax2.bar(fy_data['Financial Year'], fy_data['Capital Gain'],
                                  color=['#59a14f' if x >= 0 else '#e15759' for x in fy_data['Capital Gain']])
                    ax2.set_xlabel('Financial Year')
                    ax2.set_ylabel('Capital Gain/Loss ($)')
                    ax2.set_title('Capital Gain/Loss by Financial Year', fontweight='bold')
                    
                    # Add value labels on top of bars
                    for bar in bars:
                        height = bar.get_height()
                        value = height if height >= 0 else height
                        if height >= 0:
                            ax2.text(bar.get_x() + bar.get_width()/2., height + (0.01 * abs(height)),
                                    f'${value:,.2f}', ha='center', va='bottom', fontsize=9)
                        else:
                            ax2.text(bar.get_x() + bar.get_width()/2., height - (0.01 * abs(height)),
                                    f'${value:,.2f}', ha='center', va='top', fontsize=9)
                    
                    # Adjust layout
                    plt.tight_layout()
                    
                    st.pyplot(fig2)
                else:
                    st.info("Financial Year data not available. Please ensure your data includes dates.")
            
            with viz_tab2:
                # New visualization - Stock Performance as bar chart
                fig3, ax3 = plt.subplots()
                
                # Create a DataFrame for comparison
                performance_df = pd.DataFrame({
                    'Date': sell_data['Sell Date'],
                    'Vest Price': sell_data['FMV at Vesting'],
                    'Sale Price': sell_data['Sale Price']
                }).sort_values('Date')
                
                # Set width of bars
                barWidth = 0.4
                
                # Set position of bars on X axis
                r1 = range(len(performance_df))
                r2 = [x + barWidth for x in r1]
                
                # Create bars
                ax3.bar(r1, performance_df['Vest Price'], width=barWidth, label='FMV at Vesting', color='#4e79a7')
                ax3.bar(r2, performance_df['Sale Price'], width=barWidth, label='Sale Price', color='#f28e2c')
                
                # Add labels and title
                ax3.set_xlabel('Sale Date')
                ax3.set_ylabel('Price ($)')
                ax3.set_title('Stock Performance: Vest Price vs Sale Price', fontweight='bold')
                ax3.set_xticks([r + barWidth/2 for r in range(len(performance_df))])
                ax3.set_xticklabels([d.strftime('%Y-%m-%d') if pd.notnull(d) else '' for d in performance_df['Date']], rotation=45)


                # Add value labels on top of bars
                for i, v in enumerate(performance_df['Vest Price']):
                    ax3.text(i, v + 0.1, f"${v:,.2f}", ha='center', va='bottom', fontsize=9)
                
                for i, v in enumerate(performance_df['Sale Price']):
                    ax3.text(i + barWidth, v + 0.1, f"${v:,.2f}", ha='center', va='bottom', fontsize=9)
                
                # Add legend with better positioning
                ax3.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=2)
                
                # Adjust layout to make room for legend and rotated x-labels
                plt.tight_layout(rect=[0, 0.1, 1, 0.95])
                
                st.pyplot(fig3)
            
            with viz_tab3:
                # New visualization - Capital Gains vs Tax by Financial Year
                if "Financial Year" in sell_data.columns:
                    # Group by Financial Year
                    fy_data = sell_data.groupby("Financial Year")[["Capital Gain", "Tax on CG"]].sum().reset_index()
                    
                    # Create bar chart
                    fig4, ax4 = plt.subplots()
                    
                    # Set width of bars
                    barWidth = 0.35
                    
                    # Set position of bars on X axis
                    r1 = range(len(fy_data))
                    r2 = [x + barWidth for x in r1]
                    
                    # Create bars with theme colors
                    ax4.bar(r1, fy_data['Capital Gain'], width=barWidth, label='Capital Gain', color='#4e79a7')
                    ax4.bar(r2, fy_data['Tax on CG'], width=barWidth, label='Tax on CG', color='#e15759')
                    
                    # Add labels and title
                    ax4.set_xlabel('Financial Year')
                    ax4.set_ylabel('Amount ($)')
                    ax4.set_title('Capital Gains vs Tax by Financial Year', fontweight='bold')
                    ax4.set_xticks([r + barWidth/2 for r in range(len(fy_data))])
                    ax4.set_xticklabels(fy_data['Financial Year'])
                    
                    # Add value labels on top of bars
                    for i, v in enumerate(fy_data['Capital Gain']):
                        ax4.text(i, v + 0.1, f"${v:,.2f}", ha='center', fontsize=9)
                    
                    for i, v in enumerate(fy_data['Tax on CG']):
                        ax4.text(i + barWidth, v + 0.1, f"${v:,.2f}", ha='center', fontsize=9)
                    
                    # Add legend with better positioning
                    ax4.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=2)
                    
                    # Adjust layout to make room for legend
                    plt.tight_layout(rect=[0, 0.05, 1, 0.95])
                    
                    st.pyplot(fig4)
                else:
                    st.info("Financial Year data not available. Please ensure your data includes dates.")
        
        # Export to Excel
        rsu_sales_buffer = BytesIO()
        with pd.ExcelWriter(rsu_sales_buffer, engine='openpyxl') as writer:
            sell_data.to_excel(writer, index=False, sheet_name="RSU Sales")
        rsu_sales_buffer.seek(0)
        
        st.download_button("📥 Download RSU Sales (Excel)",
                          data=rsu_sales_buffer,
                          file_name="rsu_sales.xlsx",
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Add button to clear all data
        if st.button("🗑️ Clear All Sales Data"):
            if st.session_state.get("confirm_clear_sales", False):
                st.session_state["sales_data"] = pd.DataFrame()
                st.rerun()
            else:
                st.session_state["confirm_clear_sales"] = True
                st.warning("Click again to confirm clearing all sales data.")
    elif sales_file:
        sell_data = pd.read_excel(sales_file)
        
        # Process sales data
        sell_data = tax_engine.process_sales_data(sell_data, tax_rate)
        
        # No "Add New Sales Entry" button - users can add rows directly in the data editor
        
        # Display formatted data (editable)
        edited_sell_data = st.data_editor(
            sell_data.style.format({
                "Capital Gain": "${:,.2f}",
                "Tax on CG": "${:,.2f}",
                "Net Proceeds": "${:,.2f}",
                "Marginal Tax Rate": "{:.1f}%"
            }),
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key="sales_data_editor"  # Add a key for the data editor
        )
        
        # Process the edited data only if it has changed
        if edited_sell_data is not None and not edited_sell_data.equals(sell_data):
            # Process the edited data with tax engine
            processed_sell_data = tax_engine.process_sales_data(edited_sell_data, tax_rate)
            st.session_state["sales_data"] = processed_sell_data
            sell_data = processed_sell_data
            
            # Force rerun to update the UI immediately with recalculated values
            st.rerun()
        
        # Create visualizations
        if not sell_data.empty:
            st.markdown("### 📊 Visualizations")
            
            # Create tabs for different visualizations
            viz_tab1, viz_tab2, viz_tab3 = st.tabs([
                "Capital Gain/Loss by Financial Year",
                "Stock Performance",
                "Capital Gains vs Tax"
            ])
            
            with viz_tab2:
                # New visualization - Capital Gain/Loss by Financial Year
                if "Financial Year" in sell_data.columns:
                    # Group by Financial Year
                    fy_data = sell_data.groupby("Financial Year")["Capital Gain"].sum().reset_index()
                    
                    # Create bar chart
                    fig2, ax2 = plt.subplots(figsize=(10, 6))
                    bars = ax2.bar(fy_data['Financial Year'], fy_data['Capital Gain'], color=['green' if x >= 0 else 'red' for x in fy_data['Capital Gain']])
                    ax2.set_xlabel('Financial Year')
                    ax2.set_ylabel('Capital Gain/Loss ($)')
                    ax2.set_title('Capital Gain/Loss by Financial Year')
                    
                    # Add value labels on top of bars
                    for bar in bars:
                        height = bar.get_height()
                        value = height if height >= 0 else height
                        ax2.text(bar.get_x() + bar.get_width()/2., height + (0.01 * abs(height)),
                                f'${value:,.2f}', ha='center', va='bottom', rotation=0)
                    
                    st.pyplot(fig2)
                else:
                    st.info("Financial Year data not available. Please ensure your data includes dates.")
            
            with viz_tab2:
                # New visualization - Stock Performance as bar chart
                fig3, ax3 = plt.subplots()
                
                # Create a DataFrame for comparison
                performance_df = pd.DataFrame({
                    'Date': sell_data['Sell Date'],
                    'Vest Price': sell_data['FMV at Vesting'],
                    'Sale Price': sell_data['Sale Price']
                }).sort_values('Date')
                
                # Set width of bars
                barWidth = 0.4
                
                # Set position of bars on X axis
                r1 = range(len(performance_df))
                r2 = [x + barWidth for x in r1]
                
                # Create bars
                ax3.bar(r1, performance_df['Vest Price'], width=barWidth, label='FMV at Vesting', color='#4e79a7')
                ax3.bar(r2, performance_df['Sale Price'], width=barWidth, label='Sale Price', color='#f28e2c')
                
                # Add labels and title
                ax3.set_xlabel('Sale Date')
                ax3.set_ylabel('Price ($)')
                ax3.set_title('Stock Performance: Vest Price vs Sale Price', fontweight='bold')
                ax3.set_xticks([r + barWidth/2 for r in range(len(performance_df))])
                ax3.set_xticklabels([d.strftime('%Y-%m-%d') if pd.notnull(d) else '' for d in performance_df['Date']], rotation=45)
                
                # Add value labels on top of bars
                for i, v in enumerate(performance_df['Vest Price']):
                    ax3.text(i, v + 0.1, f"${v:,.2f}", ha='center', va='bottom', fontsize=9)
                
                for i, v in enumerate(performance_df['Sale Price']):
                    ax3.text(i + barWidth, v + 0.1, f"${v:,.2f}", ha='center', va='bottom', fontsize=9)
                
                # Add legend with better positioning
                ax3.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=2)
                
                # Adjust layout to make room for legend and rotated x-labels
                plt.tight_layout(rect=[0, 0.1, 1, 0.95])
                
                st.pyplot(fig3)
            
            with viz_tab3:
                # New visualization - Capital Gains vs Tax by Financial Year
                if "Financial Year" in sell_data.columns:
                    # Group by Financial Year
                    fy_data = sell_data.groupby("Financial Year")[["Capital Gain", "Tax on CG"]].sum().reset_index()
                    
                    # Create bar chart
                    fig4, ax4 = plt.subplots(figsize=(10, 6))
                    
                    # Set width of bars
                    barWidth = 0.35
                    
                    # Set position of bars on X axis
                    r1 = range(len(fy_data))
                    r2 = [x + barWidth for x in r1]
                    
                    # Create bars
                    ax4.bar(r1, fy_data['Capital Gain'], width=barWidth, label='Capital Gain', color='skyblue')
                    ax4.bar(r2, fy_data['Tax on CG'], width=barWidth, label='Tax on CG', color='lightcoral')
                    
                    # Add labels and title
                    ax4.set_xlabel('Financial Year')
                    ax4.set_ylabel('Amount ($)')
                    ax4.set_title('Capital Gains vs Tax by Financial Year')
                    ax4.set_xticks([r + barWidth/2 for r in range(len(fy_data))])
                    ax4.set_xticklabels(fy_data['Financial Year'])
                    
                    # Add value labels on top of bars
                    for i, v in enumerate(fy_data['Capital Gain']):
                        ax4.text(i, v + 0.1, f"${v:,.2f}", ha='center', fontsize=8)
                    
                    for i, v in enumerate(fy_data['Tax on CG']):
                        ax4.text(i + barWidth, v + 0.1, f"${v:,.2f}", ha='center', fontsize=8)
                    
                    # Add legend
                    ax4.legend()
                    
                    st.pyplot(fig4)
                else:
                    st.info("Financial Year data not available. Please ensure your data includes dates.")
        
        # Export to Excel
        rsu_sales_buffer = BytesIO()
        with pd.ExcelWriter(rsu_sales_buffer, engine='openpyxl') as writer:
            sell_data.to_excel(writer, index=False, sheet_name="RSU Sales")
        rsu_sales_buffer.seek(0)
        
        st.download_button("📥 Download RSU Sales (Excel)",
                          data=rsu_sales_buffer,
                          file_name="rsu_sales.xlsx",
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Add button to clear all data
        if st.button("🗑️ Clear All Sales Data"):
            if st.session_state.get("confirm_clear_sales", False):
                st.session_state["sales_data"] = pd.DataFrame()
                st.rerun()
            else:
                st.session_state["confirm_clear_sales"] = True
                st.warning("Click again to confirm clearing all sales data.")
    else:
        # Create an empty DataFrame with the necessary columns for sales data
        empty_sales_df = pd.DataFrame({
            "Sell Date": [],
            "Shares Sold": [],
            "Sale Price": [],
            "FMV at Vesting": [],
            "Held > 12 Months": [],
            "Marginal Tax Rate": [],
            "Gross Value": [],
            "Capital Gain": [],
            "Tax on CG": [],
            "Net Proceeds": []
        })
        
        # Process the empty DataFrame to ensure it has all required columns
        empty_sales_df = tax_engine.process_sales_data(empty_sales_df, tax_rate)
        
        # Display an editable data grid even when no data is available
        st.info("No sales data available. You can add entries directly in the grid below.")
        
        # Display formatted data (editable)
        edited_sell_data = st.data_editor(
            empty_sales_df.style.format({
                "Capital Gain": "${:,.2f}",
                "Tax on CG": "${:,.2f}",
                "Net Proceeds": "${:,.2f}",
                "Marginal Tax Rate": "{:.1f}%"
            }),
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key="sales_data_editor_empty"
        )
        
        # Process the edited data only if it has changed and is not empty
        if edited_sell_data is not None and not edited_sell_data.empty and not edited_sell_data.equals(empty_sales_df):
            # Process the edited data with tax engine
            processed_sell_data = tax_engine.process_sales_data(edited_sell_data, tax_rate)
            st.session_state["sales_data"] = processed_sell_data
            
            # Force rerun to update the UI immediately with recalculated values
            st.rerun()

# -------- TAB 3: SUMMARY --------
with tab3:
    st.subheader("📊 Summary")

    if 'schedule' in locals() and not schedule.empty:
        total_gross = schedule["Gross Value"].sum()
        total_tax = schedule["Tax Payable"].sum()
        total_net = schedule["Net Value"].sum()

        # Create columns for displaying totals in a single row
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Gross Value", f"${total_gross:,.2f}")
        with col2:
            st.metric("Total Tax Payable", f"${total_tax:,.2f}")
        with col3:
            st.metric("Total Net Value", f"${total_net:,.2f}")

        # Group by Financial Year
        grouped_fy = schedule.groupby("Financial Year")[["Gross Value", "Tax Payable", "Net Value"]].sum().reset_index()
        st.markdown("### 📅 Financial Year-wise Tax Summary")
        st.dataframe(grouped_fy.style.format({
            "Gross Value": "${:,.2f}",
            "Tax Payable": "${:,.2f}",
            "Net Value": "${:,.2f}"
        }))
    else:
        st.info("No vesting data available. Please upload or load sample vesting data to see summary.")

    if 'sell_data' in locals() and not sell_data.empty:
        total_gain = sell_data["Capital Gain"].sum()
        total_cgt_tax = sell_data["Tax on CG"].sum()
        total_net_sale = sell_data["Net Proceeds"].sum()

        st.markdown("### 💰 Capital Gains Summary")
        # Create columns for displaying capital gains totals in a single row
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Capital Gains", f"${total_gain:,.2f}")
        with col2:
            st.metric("Tax on Capital Gains", f"${total_cgt_tax:,.2f}")
        with col3:
            st.metric("Net Proceeds from RSU Sales", f"${total_net_sale:,.2f}")
        
        # Group by Financial Year
        if "Financial Year" in sell_data.columns:
            grouped_cg_fy = sell_data.groupby("Financial Year")[["Capital Gain", "Tax on CG", "Net Proceeds"]].sum().reset_index()
            st.markdown("### 📅 Financial Year-wise Capital Gain Summary")
            st.dataframe(grouped_cg_fy.style.format({
                "Capital Gain": "${:,.2f}",
                "Tax on CG": "${:,.2f}",
                "Net Proceeds": "${:,.2f}"
            }))

        if st.button("� Export Summary to PDF"):
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt="RSU Management Summary", ln=True, align="C")
            pdf.ln(10)
            pdf.cell(200, 10, txt=f"Company: {company}", ln=True)
            
            # Add Financial Year-wise Capital Gain Summary to PDF if available
            if "Financial Year" in sell_data.columns and not grouped_cg_fy.empty:
                pdf.ln(5)
                pdf.cell(200, 10, txt="Financial Year-wise Capital Gain Summary:", ln=True)
                for _, row in grouped_cg_fy.iterrows():
                    pdf.cell(200, 10, txt=f"FY {row['Financial Year']}: Capital Gain ${row['Capital Gain']:,.2f}, Tax ${row['Tax on CG']:,.2f}, Net ${row['Net Proceeds']:,.2f}", ln=True)
            pdf.cell(200, 10, txt=f"Total Gross Value: ${total_gross:,.2f}", ln=True)
            pdf.cell(200, 10, txt=f"Total Tax Payable: ${total_tax:,.2f}", ln=True)
            pdf.cell(200, 10, txt=f"Total Net Value: ${total_net:,.2f}", ln=True)
            pdf.ln(5)
            pdf.cell(200, 10, txt="Capital Gains Summary:", ln=True)
            pdf.cell(200, 10, txt=f"Total Capital Gains: ${total_gain:,.2f}", ln=True)
            pdf.cell(200, 10, txt=f"Tax on Capital Gains: ${total_cgt_tax:,.2f}", ln=True)
            pdf.cell(200, 10, txt=f"Net Proceeds: ${total_net_sale:,.2f}", ln=True)

            pdf_output = pdf.output(dest='S').encode('latin1')
            buffer = BytesIO(pdf_output)
            st.download_button("📥 Download PDF Summary", data=buffer, file_name="rsu_summary.pdf", mime="application/pdf")
    elif 'schedule' in locals() and not schedule.empty:
        st.info("No sales data available. Please upload or load sample sales data to see capital gains summary.")

# -------- TAB 4: TAX RETURN ASSISTANT --------
with tab4:
    st.subheader("📝 EOY Tax Return Assistant")
    
    # Financial year selection
    available_years = []
    
    # Get available financial years from vesting and sales data
    if 'schedule' in locals() and not schedule.empty:
        if "Financial Year" in schedule.columns:
            schedule_years = schedule["Financial Year"].unique().tolist()
            available_years.extend(schedule_years)
    
    if 'sell_data' in locals() and not sell_data.empty:
        # Ensure Financial Year column exists in sell_data
        if "Financial Year" not in sell_data.columns and "Sell Date" in sell_data.columns:
            # Use the tax_engine function for consistency
            sell_data["Financial Year"] = sell_data["Sell Date"].apply(
                lambda x: tax_engine.determine_financial_year(x) if pd.notnull(x) else "Unknown"
            )
        
        if "Financial Year" in sell_data.columns:
            sales_years = sell_data["Financial Year"].unique().tolist()
            available_years.extend(sales_years)
    
    # Remove duplicates and sort (filter out None values)
    available_years = [year for year in available_years if year is not None]
    if available_years:
        available_years = sorted(list(set(available_years)))
    else:
        available_years = []
    
    if not available_years:
        available_years = [f"{datetime.date.today().year-1}-{datetime.date.today().year}"]
    
    selected_year = st.selectbox("Select Financial Year", available_years, index=0)
    
    # Create columns for layout
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("### 📊 Income Summary")
        
        # Generate tax summary for selected year
        if 'schedule' in locals() and 'sell_data' in locals():
            # Generate simplified financial year summary
            tax_summary = tax_engine.generate_financial_year_summary(
                schedule,
                sell_data,
                selected_year,
                other_income=0.0,
                has_private_health=False,
                family=False,
                has_help_debt=False
            )
            
            # Display income summary (RSU-focused)
            income_data = {
                "Category": ["RSU Vesting Income", "Capital Gains", "CGT Discount Applied", "Net Capital Gain", "Total Taxable Income"],
                "Amount": [
                    tax_summary.ordinary_income,
                    tax_summary.capital_gains,
                    -tax_summary.cgt_discount,  # Negative as it's a reduction
                    tax_summary.net_capital_gain,
                    tax_summary.calculate_total_income()
                ]
            }
            
            income_df = pd.DataFrame(income_data)
            st.dataframe(income_df.style.format({
                "Amount": "${:,.2f}"
            }), use_container_width=True)
            
            st.markdown("### 💰 Tax Summary")
            
            # Display simplified tax summary
            tax_data = {
                "Category": [
                    "Estimated Tax on Income",
                    "Tax Already Withheld",
                    "Remaining Tax Payable"
                ],
                "Amount": [
                    tax_summary.estimated_tax,
                    -tax_summary.tax_withheld,  # Negative as it reduces payable
                    tax_summary.get_remaining_tax_payable()
                ]
            }
            
            tax_df = pd.DataFrame(tax_data)
            st.dataframe(tax_df.style.format({
                "Amount": "${:,.2f}"
            }), use_container_width=True)
        else:
            st.info("Please load or enter RSU vesting and sales data to generate tax summaries.")
    
    with col2:
        st.markdown("### 📝 ATO Tax Return Information")
        
        if 'tax_summary' in locals():
            # Map to ATO items
            ato_items = tax_summary.map_to_ato_items()
            
            # Display ATO item codes
            for code, amount in ato_items.items():
                if amount > 0:  # Only show non-zero items
                    st.write(f"**{code}:** ${amount:,.2f}")
            
            st.markdown("### 📋 Tax Return Checklist")
            
            # Create a simplified checklist of documents needed
            documents = [
                "Employee Payment Summary",
                "Employee Share Scheme (ESS) Statement",
                "Share Sale Contract Notes"
            ]
            
            for doc in documents:
                st.checkbox(doc, key=f"doc_{doc}")
        else:
            st.info("Tax return information will appear here once data is loaded.")
    
    # Export options
    if 'tax_summary' in locals():
        st.markdown("### 📥 Export Options")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("📊 Export Tax Summary (Excel)"):
                # Create Excel with tax summary
                tax_buffer = BytesIO()
                with pd.ExcelWriter(tax_buffer, engine='openpyxl') as writer:
                    # Income sheet
                    income_df.to_excel(writer, index=False, sheet_name="Income Summary")
                    
                    # Tax sheet
                    tax_df.to_excel(writer, index=False, sheet_name="Tax Summary")
                    
                    # ATO Items sheet
                    ato_df = pd.DataFrame({
                        "Item Code": list(ato_items.keys()),
                        "Amount": list(ato_items.values())
                    })
                    ato_df.to_excel(writer, index=False, sheet_name="ATO Items")
                    
                tax_buffer.seek(0)
                
                st.download_button(
                    "📥 Download Tax Summary",
                    data=tax_buffer,
                    file_name=f"tax_summary_{selected_year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col2:
            if st.button("📄 Export ATO Format (PDF)"):
                # Create PDF with ATO format
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", size=12)
                pdf.cell(200, 10, txt=f"Tax Return Summary - {selected_year}", ln=True, align="C")
                pdf.ln(10)
                
                # Company and personal info
                pdf.cell(200, 10, txt=f"Company: {company}", ln=True)
                pdf.cell(200, 10, txt=f"Tax Residency: {tax_residency}", ln=True)
                pdf.ln(5)
                
                # ATO Items
                pdf.set_font("Arial", 'B', size=12)
                pdf.cell(200, 10, txt="ATO Tax Return Items:", ln=True)
                pdf.set_font("Arial", size=12)
                
                for code, amount in ato_items.items():
                    if amount > 0:  # Only show non-zero items
                        pdf.cell(200, 10, txt=f"{code}: ${amount:,.2f}", ln=True)
                
                pdf_output = pdf.output(dest='S').encode('latin1')
                buffer = BytesIO(pdf_output)
                
                st.download_button(
                    "📥 Download ATO Format",
                    data=buffer,
                    file_name=f"ato_tax_return_{selected_year}.pdf",
                    mime="application/pdf"
                )
        
        with col3:
            if st.button("📋 Export Tax Documentation Checklist"):
                # Create PDF with checklist
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", size=12)
                pdf.cell(200, 10, txt="RSU Tax Documentation Checklist", ln=True, align="C")
                pdf.ln(10)
                
                # Checklist items
                pdf.set_font("Arial", 'B', size=12)
                pdf.cell(200, 10, txt="Required Documents:", ln=True)
                pdf.set_font("Arial", size=12)
                
                for doc in documents:
                    pdf.cell(200, 10, txt=f"□ {doc}", ln=True)
                
                pdf.ln(5)
                pdf.cell(200, 10, txt="Notes:", ln=True)
                pdf.ln(20)  # Space for notes
                
                # Tax deadlines
                pdf.set_font("Arial", 'B', size=12)
                pdf.cell(200, 10, txt="Important Tax Deadlines:", ln=True)
                pdf.set_font("Arial", size=12)
                pdf.cell(200, 10, txt=f"• Individual Tax Return Due: 31 October {selected_year.split('-')[1]}", ln=True)
                
                pdf_output = pdf.output(dest='S').encode('latin1')
                buffer = BytesIO(pdf_output)
                
                st.download_button(
                    "📥 Download Checklist",
                    data=buffer,
                    file_name="tax_documentation_checklist.pdf",
                    mime="application/pdf"
                )

# -------- TAB 5: OPTIMIZATION --------
with tab5:
    st.subheader("💡 Tax Optimization Engine")
    st.info("Tax Optimization Engine will be implemented in the next phase.")
